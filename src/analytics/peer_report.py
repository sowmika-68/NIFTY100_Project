"""Generate peer comparison Excel report with percentile coloring and benchmarks.

Creates `output/peer_comparison.xlsx` with one sheet per peer group.
Each sheet contains: company_id, company_name, 20 metrics, and percentile rank columns.
Percentile color-coding: green >=75, yellow 25-75, red <=25. Benchmark row highlighted.
Summary median row appended at bottom.
"""

import os
import sqlite3
from typing import List

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

from src.screener.engine import ScreenerEngine

OUT_PATH = "output/peer_comparison.xlsx"
PEER_XLSX = "data/peer_groups.xlsx"

# 20 metric columns
METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "cfo_pat_ratio",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "sales",
    "net_profit",
    "composite_quality_score",
]


def percent_rank_vals(s: pd.Series) -> pd.Series:
    # Using (rank-1)/(n-1) scaled to 0-100
    s = pd.to_numeric(s, errors="coerce")
    n = s.dropna().shape[0]
    if n == 0:
        return pd.Series([pd.NA] * len(s), index=s.index)
    if n == 1:
        out = pd.Series(100.0, index=s.index)
        out[s.isna()] = pd.NA
        return out
    ranks = s.rank(method="min", na_option="keep")
    pr = (ranks - 1) / (n - 1) * 100.0
    pr = pr.where(~s.isna(), other=pd.NA)
    return pr


def create_report(peer_xlsx: str = PEER_XLSX, out_path: str = OUT_PATH):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    if not os.path.exists(peer_xlsx):
        raise FileNotFoundError(f"Peer groups file not found: {peer_xlsx}")

    df_peers = pd.read_excel(peer_xlsx)
    if "company_id" not in df_peers.columns or "peer_group_name" not in df_peers.columns:
        raise ValueError("peer_groups.xlsx must contain 'company_id' and 'peer_group_name' columns")

    # Prepare screener data
    engine = ScreenerEngine()
    engine.prepare()
    df = engine.df.copy()

    # Merge peer group info and benchmark flag
    df = df.merge(df_peers[["company_id", "peer_group_name", "is_benchmark"]], on="company_id", how="left")

    # Only peer groups that exist in the peer_xlsx
    peer_groups = df_peers["peer_group_name"].unique().tolist()

    # Build Excel writer
    # prepare per-sheet data and medians so missing columns are handled
    sheet_data = {}
    sheet_medians = {}

    for pg in peer_groups:
        group = df[df["peer_group_name"] == pg].copy()
        sheet_name = str(pg)[:31]
        if group.empty:
            cols = ["company_id", "company_name"] + METRICS + [m + "_percentile" for m in METRICS]
            sheet_data[sheet_name] = pd.DataFrame(columns=cols)
            sheet_medians[sheet_name] = ([], [])
            continue

        # Build base frame with columns present or NaN
        data = {"company_id": group["company_id"].values, "company_name": group["company_name"].values}
        for m in METRICS:
            data[m] = group[m] if m in group.columns else pd.Series([pd.NA] * len(group), index=group.index)

        out = pd.DataFrame(data, index=group.index)

        # Compute percentiles per metric within group
        median_vals = []
        median_pr = []
        for m in METRICS:
            if m not in group.columns:
                out[m + "_percentile"] = pd.NA
                median_vals.append(pd.NA)
                median_pr.append(pd.NA)
                continue

            pr = percent_rank_vals(group[m])
            if m == "debt_to_equity":
                pr = pr.apply(lambda x: 100.0 - x if pd.notna(x) else pd.NA)
            out[m + "_percentile"] = pr.values

            # medians
            median_vals.append(np.nanmedian(pd.to_numeric(out[m], errors="coerce")))
            median_pr.append(np.nanmedian(pd.to_numeric(out[m + "_percentile"], errors="coerce")))

        sheet_data[sheet_name] = out
        sheet_medians[sheet_name] = (median_vals, median_pr)

    # write all sheets
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df_out in sheet_data.items():
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    # Now open with openpyxl to apply formatting and append median rows and highlight benchmarks
    wb = load_workbook(out_path)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    for pg in peer_groups:
        sheet_name = str(pg)[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # find header row and column indices
        header = [cell.value for cell in ws[1]]
        # determine percentile columns positions
        perc_cols = [i + 1 for i, h in enumerate(header) if h and h.endswith("_percentile")]

        # apply conditional formatting to percentile columns
        for col_idx in perc_cols:
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            start_row = 2
            end_row = ws.max_row
            # green >=75
            ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}",
                                          CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=green))
            # yellow between 25 and 75
            ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}",
                                          CellIsRule(operator="between", formula=["25", "75"], fill=yellow))
            # red <=25
            ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}",
                                          CellIsRule(operator="lessThanOrEqual", formula=["25"], fill=red))

        # highlight benchmark row(s)
        # find company_ids that are benchmark in peer_xlsx
        # load peer_xlsx
        df_peers = pd.read_excel(peer_xlsx)
        benchmarks = df_peers[(df_peers["peer_group_name"] == pg) & (df_peers.get("is_benchmark") == True)]["company_id"].tolist()
        if benchmarks:
            # iterate rows and highlight matches
            for r in range(2, ws.max_row + 1):
                cid = ws.cell(row=r, column=1).value
                if cid in benchmarks:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = gold

        # append median row at bottom
        median_row = ["MEDIAN", ""]
        for v in median_vals:
            median_row.append(None if pd.isna(v) else v)
        for v in median_pr:
            median_row.append(None if pd.isna(v) else v)

        # append
        ws.append(median_row)
        # make median row bold? skipped to keep simple

    wb.save(out_path)
    print(f"Peer comparison Excel written to: {out_path}")


if __name__ == "__main__":
    create_report()
