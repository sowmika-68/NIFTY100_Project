"""
Sprint 3 - Day 15
Financial Screener Engine
Part 1/4
"""

import os
import sqlite3
import yaml
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

DATABASE = "db/nifty100.db"
CONFIG = "config/screener_config.yaml"


class ScreenerEngine:

    def __init__(self):

        self.connection = sqlite3.connect(DATABASE)

        self.config = self.load_config()

        self.df = pd.DataFrame()
        self.is_prepared = False

    def load_config(self):
        """
        Load YAML configuration
        """

        with open(CONFIG, "r") as file:

            config = yaml.safe_load(file)

        return config

    def table_has_column(self, table, column):
        cursor = self.connection.cursor()
        cursor.execute(f"PRAGMA table_info({table})")
        return any(row[1] == column for row in cursor.fetchall())

    def load_data(self):
        """
        Load all required data from SQLite
        """

        optional_return_on_capital = ""
        if self.table_has_column("financial_ratios", "return_on_capital_employed_pct"):
            optional_return_on_capital = "fr.return_on_capital_employed_pct,\n\n"

        query = f"""
        SELECT

            fr.company_id,
            fr.year,

            c.company_name,

            s.broad_sector,
            s.sub_sector,

            fr.return_on_equity_pct,

            {optional_return_on_capital}
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,

            fr.debt_to_equity,
            fr.interest_coverage,
            fr.asset_turnover,

            fr.free_cash_flow_cr,

            fr.earnings_per_share,

            fr.dividend_payout_ratio_pct,

            fr.book_value_per_share,

            fr.total_debt_cr,

            fr.cash_from_operations_cr,

            an.compounded_sales_growth_5yr,
            an.compounded_sales_growth_3yr,
            an.compounded_profit_growth_5yr,
            an.compounded_profit_growth_3yr,

            mc.market_cap_crore,

            mc.enterprise_value_crore,

            mc.pe_ratio,

            mc.pb_ratio,

            mc.ev_ebitda,

            mc.dividend_yield_pct,

            pl.sales,

            pl.net_profit,

            pl.opm_percentage

        FROM financial_ratios fr

        LEFT JOIN profitandloss pl

        ON fr.company_id = pl.company_id
        AND fr.year = pl.year

        LEFT JOIN companies c

        ON fr.company_id = c.id

        LEFT JOIN sectors s

        ON fr.company_id = s.company_id

        LEFT JOIN market_cap mc

        ON fr.company_id = mc.company_id

        LEFT JOIN (
            SELECT
                company_id,
                MAX(CASE WHEN compounded_sales_growth LIKE '%5 Years:%' THEN compounded_sales_growth END) AS compounded_sales_growth_5yr,
                MAX(CASE WHEN compounded_sales_growth LIKE '%3 Years:%' THEN compounded_sales_growth END) AS compounded_sales_growth_3yr,
                MAX(CASE WHEN compounded_profit_growth LIKE '%5 Years:%' THEN compounded_profit_growth END) AS compounded_profit_growth_5yr,
                MAX(CASE WHEN compounded_profit_growth LIKE '%3 Years:%' THEN compounded_profit_growth END) AS compounded_profit_growth_3yr
            FROM analysis
            GROUP BY company_id
        ) an

        ON fr.company_id = an.company_id

        """

        self.df = pd.read_sql_query(
            query,
            self.connection
        )

        print("=" * 60)
        print("Master Data Loaded")
        print(self.df.shape)
        print("=" * 60)

    @staticmethod
    def extract_timeframe_percent(series, timeframe):
        pattern = rf"{timeframe}[^0-9+-]*([+-]?[0-9]*\.?[0-9]+)%"
        values = series.astype(str)
        extracted = values.str.extract(pattern, expand=False)
        if extracted is None:
            extracted = pd.Series([None] * len(series), index=series.index)
        fallback = values.str.extract(r"([+-]?[0-9]*\.?[0-9]+)%", expand=False)
        return pd.to_numeric(extracted.fillna(fallback), errors="coerce")

    @staticmethod
    def calculate_cagr(start_value, end_value, years):
        if pd.isna(start_value) or pd.isna(end_value):
            return np.nan
        if years <= 0:
            return np.nan
        if start_value <= 0 or end_value <= 0:
            return np.nan
        try:
            return ((end_value / start_value) ** (1 / years) - 1) * 100
        except Exception:
            return np.nan

    @staticmethod
    def calculate_period_cagr_series(values, dates, years):
        dates = pd.to_datetime(dates, errors="coerce")
        result = np.full(len(values), np.nan)
        values = pd.to_numeric(values, errors="coerce")
        for idx in range(len(values)):
            current_date = dates.iat[idx]
            if pd.isna(current_date):
                continue
            start_cutoff = current_date - pd.DateOffset(years=years)
            earlier_mask = dates <= start_cutoff
            earlier_indices = np.flatnonzero(np.asarray(earlier_mask))
            if len(earlier_indices) == 0:
                continue
            start_value = values.iloc[earlier_indices[-1]]
            end_value = values.iat[idx]
            result[idx] = ScreenerEngine.calculate_cagr(start_value, end_value, years)
        return result

    def compute_derived_metrics(self):
        self.df["year_date"] = pd.to_datetime(
            self.df["year"],
            format="%b %Y",
            errors="coerce"
        )
        self.df["debt_to_equity"] = pd.to_numeric(
            self.df["debt_to_equity"],
            errors="coerce"
        )
        self.df["sales"] = pd.to_numeric(self.df["sales"], errors="coerce")
        self.df["net_profit"] = pd.to_numeric(self.df["net_profit"], errors="coerce")
        self.df["earnings_per_share"] = pd.to_numeric(
            self.df["earnings_per_share"],
            errors="coerce"
        )

        self.df = self.df.sort_values(["company_id", "year_date"])

        self.df["revenue_cagr_5yr"] = np.nan
        self.df["revenue_cagr_3yr"] = np.nan
        self.df["pat_cagr_5yr"] = np.nan
        self.df["pat_cagr_3yr"] = np.nan
        self.df["eps_cagr_5yr"] = np.nan
        self.df["fcf_cagr_5yr"] = np.nan

        for _, group in self.df.groupby("company_id", sort=False):
            index = group.index
            self.df.loc[index, "revenue_cagr_5yr"] = self.calculate_period_cagr_series(
                group["sales"], group["year_date"], 5
            )
            self.df.loc[index, "revenue_cagr_3yr"] = self.calculate_period_cagr_series(
                group["sales"], group["year_date"], 3
            )
            self.df.loc[index, "pat_cagr_5yr"] = self.calculate_period_cagr_series(
                group["net_profit"], group["year_date"], 5
            )
            self.df.loc[index, "pat_cagr_3yr"] = self.calculate_period_cagr_series(
                group["net_profit"], group["year_date"], 3
            )
            self.df.loc[index, "eps_cagr_5yr"] = self.calculate_period_cagr_series(
                group["earnings_per_share"], group["year_date"], 5
            )
            self.df.loc[index, "fcf_cagr_5yr"] = self.calculate_period_cagr_series(
                group["free_cash_flow_cr"], group["year_date"], 5
            )

        self.df["de_ratio_prev_year"] = self.df.groupby("company_id")["debt_to_equity"].shift(1)
        self.df["de_ratio_decline"] = self.df["debt_to_equity"] < self.df["de_ratio_prev_year"]

    def latest_year_only(self):
        """
        Keep only latest financial year for every company
        """

        self.df = (
            self.df
            .sort_values("year_date")
            .groupby("company_id")
            .tail(1)
            .reset_index(drop=True)
        )

        print("Latest Year Records :", len(self.df))

    def clean_data(self):
        """
        Handle missing values
        """

        numeric_columns = self.df.select_dtypes(
            include=np.number
        ).columns

        self.df[numeric_columns] = (
            self.df[numeric_columns]
            .fillna(0)
        )

        self.df["broad_sector"] = (
            self.df["broad_sector"]
            .fillna("Unknown")
        )

        self.df["sub_sector"] = (
            self.df["sub_sector"]
            .fillna("Unknown")
        )

        for col in [
            "revenue_cagr_5yr",
            "revenue_cagr_3yr",
            "pat_cagr_5yr",
            "pat_cagr_3yr",
            "eps_cagr_5yr",
            "fcf_cagr_5yr",
        ]:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(
                    self.df[col],
                    errors="coerce"
                ).fillna(0)

    def debt_free_logic(self):
        """
        Debt Free means Infinite ICR
        """

        self.df.loc[
            self.df["interest_coverage"] == "Debt Free",
            "interest_coverage"
        ] = np.inf

        self.df["interest_coverage"] = pd.to_numeric(
            self.df["interest_coverage"],
            errors="coerce"
        ).fillna(np.inf)

    def winsorize_series(self, series):
        numeric = pd.to_numeric(
            pd.Series(series).replace([np.inf, -np.inf], np.nan),
            errors="coerce"
        )
        lower = numeric.quantile(0.10)
        upper = numeric.quantile(0.90)
        return numeric.clip(lower, upper), lower, upper

    def scale_series(self, series, lower, upper, invert=False):
        if pd.isna(lower) or pd.isna(upper):
            return pd.Series(0.0, index=series.index)

        if lower == upper:
            return pd.Series(100.0, index=series.index)

        clipped = series.clip(lower, upper)
        scaled = (clipped - lower) / (upper - lower) * 100
        if invert:
            scaled = 100 - scaled
        return scaled.fillna(0.0)

    def metric_score(self, series, invert=False):
        winsorized, lower, upper = self.winsorize_series(series)
        return self.scale_series(winsorized, lower, upper, invert=invert)

    def compute_composite_quality_score(self):
        self.df["cfo_pat_ratio"] = np.where(
            self.df["net_profit"] == 0,
            np.nan,
            self.df["cash_from_operations_cr"] / self.df["net_profit"]
        )
        self.df["free_cash_flow_positive"] = (
            self.df["free_cash_flow_cr"] > 0
        ).astype(int)

        self.df["roe_score"] = self.metric_score(self.df.get("return_on_equity_pct", pd.Series(np.nan)))
        self.df["roce_score"] = self.metric_score(
            self.df.get("return_on_capital_employed_pct", pd.Series(np.nan))
        )
        self.df["npm_score"] = self.metric_score(self.df.get("net_profit_margin_pct", pd.Series(np.nan)))
        self.df["fcf_cagr_score"] = self.metric_score(self.df.get("fcf_cagr_5yr", pd.Series(np.nan)))
        self.df["cfo_pat_score"] = self.metric_score(self.df.get("cfo_pat_ratio", pd.Series(np.nan)))
        self.df["revenue_cagr_score"] = self.metric_score(self.df.get("revenue_cagr_5yr", pd.Series(np.nan)))
        self.df["pat_cagr_score"] = self.metric_score(self.df.get("pat_cagr_5yr", pd.Series(np.nan)))
        self.df["debt_to_equity_score"] = self.metric_score(
            self.df.get("debt_to_equity", pd.Series(np.nan)),
            invert=True
        )
        self.df["interest_coverage_score"] = self.metric_score(
            self.df.get("interest_coverage", pd.Series(np.nan))
        )
        self.df["free_cash_flow_flag_score"] = self.df["free_cash_flow_positive"] * 100

        self.df["composite_quality_score_raw"] = (
            0.15 * self.df["roe_score"]
            + 0.10 * self.df["roce_score"]
            + 0.10 * self.df["npm_score"]
            + 0.15 * self.df["fcf_cagr_score"]
            + 0.10 * self.df["cfo_pat_score"]
            + 0.05 * self.df["free_cash_flow_flag_score"]
            + 0.10 * self.df["revenue_cagr_score"]
            + 0.10 * self.df["pat_cagr_score"]
            + 0.10 * self.df["debt_to_equity_score"]
            + 0.05 * self.df["interest_coverage_score"]
        )

        sector_min = self.df.groupby("broad_sector")["composite_quality_score_raw"].transform("min")
        sector_max = self.df.groupby("broad_sector")["composite_quality_score_raw"].transform("max")

        self.df["composite_quality_score"] = np.where(
            sector_max > sector_min,
            (self.df["composite_quality_score_raw"] - sector_min)
            / (sector_max - sector_min) * 100,
            100,
        )
        self.df["composite_quality_score"] = self.df["composite_quality_score"].fillna(0)

    def get_export_columns(self):
        columns = [
            "company_id",
            "company_name",
            "broad_sector",
            "sub_sector",
            "year",
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "asset_turnover",
            "free_cash_flow_cr",
            "cash_from_operations_cr",
            "fcf_cagr_5yr",
            "cfo_pat_ratio",
            "free_cash_flow_positive",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "eps_cagr_5yr",
            "market_cap_crore",
            "pe_ratio",
            "pb_ratio",
            "dividend_yield_pct",
            "dividend_payout_ratio_pct",
            "sales",
            "net_profit",
            "opm_percentage",
            "composite_quality_score",
            "threshold_pass_score",
        ]
        return [col for col in columns if col in self.df.columns]

    def generate_screener_output(self, output_path="output/screener_output.xlsx"):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        self.prepare()
        self.apply_threshold_filters()

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for preset_name in self.PRESET_DEFINITIONS:
                result = self.evaluate_preset(preset_name)
                columns = self.get_export_columns()
                df_out = result[columns].copy()
                df_out.to_excel(writer, sheet_name=preset_name[:31], index=False)
                worksheet = writer.sheets[preset_name[:31]]

                for idx, column in enumerate(df_out.columns, start=1):
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = 15

                for condition in self.PRESET_DEFINITIONS[preset_name]:
                    column = condition["column"]
                    if column not in df_out.columns:
                        continue
                    col_idx = df_out.columns.get_loc(column)
                    col_letter = get_column_letter(col_idx + 1)
                    start_row = 2
                    end_row = len(df_out) + 1
                    threshold = condition.get("threshold")
                    range_string = f"{col_letter}{start_row}:{col_letter}{end_row}"

                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    if condition["comparator"] == "min":
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="greaterThanOrEqual",
                                formula=[str(threshold)],
                                fill=green_fill,
                            ),
                        )
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="lessThan",
                                formula=[str(threshold)],
                                fill=red_fill,
                            ),
                        )
                    elif condition["comparator"] == "max":
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="lessThanOrEqual",
                                formula=[str(threshold)],
                                fill=green_fill,
                            ),
                        )
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="greaterThan",
                                formula=[str(threshold)],
                                fill=red_fill,
                            ),
                        )
                    elif condition["comparator"] == "equals":
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="equal",
                                formula=[str(threshold)],
                                fill=green_fill,
                            ),
                        )
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="notEqual",
                                formula=[str(threshold)],
                                fill=red_fill,
                            ),
                        )
                    elif condition["comparator"] == "is_true":
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="equal",
                                formula=["TRUE"],
                                fill=green_fill,
                            ),
                        )
                        worksheet.conditional_formatting.add(
                            range_string,
                            CellIsRule(
                                operator="notEqual",
                                formula=["TRUE"],
                                fill=red_fill,
                            ),
                        )

    def prepare(self):

        self.load_data()

        self.compute_derived_metrics()

        self.latest_year_only()

        self.clean_data()

        self.debt_free_logic()

        self.compute_composite_quality_score()

        self.is_prepared = True

    def resolve_filter_column(self, filter_name, settings):
        """
        Resolve the dataframe column name for a configured filter.
        """

        column_name = settings.get("column") or filter_name
        column_name = self.FILTER_COLUMN_ALIASES.get(column_name, column_name)

        if column_name not in self.df.columns:
            column_name = self.FILTER_COLUMN_ALIASES.get(filter_name, column_name)

        return column_name

    def apply_threshold_filters(self):
        """
        Apply configured threshold filters and compute the composite score.
        """

        filters = self.config.get("filters", {})

        self.df = self.df.copy()
        pass_filters = {}

        for filter_name, settings in filters.items():
            column = self.resolve_filter_column(filter_name, settings)

            if column not in self.df.columns:
                raise KeyError(
                    f"Filter '{filter_name}' expects column '{column}' but it is not available in the data."
                )

            comparator = (
                "min"
                if "min" in settings
                else "max"
                if "max" in settings
                else "equals"
                if "equals" in settings
                else None
            )
            if comparator is None:
                continue

            threshold = settings.get(comparator)
            values = pd.to_numeric(self.df[column], errors="coerce")

            if comparator == "min":
                mask = values >= threshold
            elif comparator == "max":
                mask = values <= threshold
            else:
                mask = values == threshold

            if filter_name == "debt_to_equity" and comparator == "max":
                financials = self.df["broad_sector"] == "Financials"
                mask = mask | financials

            pass_filters[filter_name] = mask.fillna(False)

        pass_df = pd.DataFrame(pass_filters, index=self.df.index)

        if pass_df.empty:
            self.df["threshold_pass_score"] = 0.0
            return

        self.df = pd.concat([self.df, pass_df.add_prefix("pass_")], axis=1)
        self.df["threshold_pass_score"] = (
            pass_df.sum(axis=1, numeric_only=True)
            / len(pass_df.columns)
        )

        self.df.sort_values(
            by=["threshold_pass_score", "composite_quality_score"],
            ascending=[False, False],
            inplace=True,
        )
        self.df.reset_index(drop=True, inplace=True)

    def screen(self):
        """
        Run the full screener flow and return the scored, sorted DataFrame.
        """

        self.prepare()
        self.apply_threshold_filters()
        return self.df

    def evaluate_preset(self, preset_name):
        if preset_name not in self.PRESET_DEFINITIONS:
            raise KeyError(f"Unknown preset: {preset_name}")

        if not self.is_prepared:
            self.prepare()
        if "threshold_pass_score" not in self.df.columns:
            self.apply_threshold_filters()

        masks = []
        for condition in self.PRESET_DEFINITIONS[preset_name]:
            masks.append(self.evaluate_condition(condition))

        if masks:
            preset_mask = pd.concat(masks, axis=1).all(axis=1)
            result = self.df.loc[preset_mask].copy()
        else:
            result = self.df.copy()

        result.sort_values(by="composite_quality_score", ascending=False, inplace=True)
        result.reset_index(drop=True, inplace=True)
        return result

    def evaluate_condition(self, condition):
        column = condition["column"]
        comparator = condition["comparator"]
        threshold = condition.get("threshold")

        if column == "de_ratio_decline":
            return self.df["de_ratio_decline"].fillna(False)

        if column not in self.df.columns:
            raise KeyError(
                f"Preset condition expects column '{column}' but it is not available in the data."
            )

        values = pd.to_numeric(self.df[column], errors="coerce")

        if comparator == "min":
            return values >= threshold
        if comparator == "max":
            return values <= threshold
        if comparator == "equals":
            return values == threshold
        if comparator == "is_true":
            return values.astype(bool) == bool(threshold)

        raise ValueError(f"Unsupported comparator: {comparator}")

    PRESET_DEFINITIONS = {
        "quality_compounder": [
            {"column": "return_on_equity_pct", "comparator": "min", "threshold": 15},
            {"column": "debt_to_equity", "comparator": "max", "threshold": 1.0},
            {"column": "free_cash_flow_cr", "comparator": "min", "threshold": 0},
            {"column": "revenue_cagr_5yr", "comparator": "min", "threshold": 10},
        ],
        "value_pick": [
            {"column": "pe_ratio", "comparator": "max", "threshold": 20},
            {"column": "pb_ratio", "comparator": "max", "threshold": 3.0},
            {"column": "debt_to_equity", "comparator": "max", "threshold": 2.0},
            {"column": "dividend_yield_pct", "comparator": "min", "threshold": 1},
        ],
        "growth_accelerator": [
            {"column": "pat_cagr_5yr", "comparator": "min", "threshold": 20},
            {"column": "revenue_cagr_5yr", "comparator": "min", "threshold": 15},
            {"column": "debt_to_equity", "comparator": "max", "threshold": 2.0},
        ],
        "dividend_champion": [
            {"column": "dividend_yield_pct", "comparator": "min", "threshold": 2},
            {"column": "dividend_payout_ratio_pct", "comparator": "max", "threshold": 80},
            {"column": "free_cash_flow_cr", "comparator": "min", "threshold": 0},
        ],
        "debt_free_blue_chip": [
            {"column": "debt_to_equity", "comparator": "equals", "threshold": 0},
            {"column": "return_on_equity_pct", "comparator": "min", "threshold": 12},
            {"column": "sales", "comparator": "min", "threshold": 5000},
        ],
        "turnaround_watch": [
            {"column": "revenue_cagr_3yr", "comparator": "min", "threshold": 10},
            {"column": "free_cash_flow_cr", "comparator": "min", "threshold": 0},
            {"column": "de_ratio_decline", "comparator": "is_true", "threshold": True},
        ],
    }

    FILTER_COLUMN_ALIASES = {
        "roe": "return_on_equity_pct",
        "debt_to_equity": "debt_to_equity",
        "free_cash_flow": "free_cash_flow_cr",
        "revenue_growth": "revenue_cagr_5yr",
        "profit_growth": "pat_cagr_5yr",
        "eps_growth": "eps_cagr_5yr",
        "revenue_cagr_5yr": "revenue_cagr_5yr",
        "pat_cagr_5yr": "pat_cagr_5yr",
        "revenue_cagr_3yr": "revenue_cagr_3yr",
        "pat_cagr_3yr": "pat_cagr_3yr",
        "operating_margin": "operating_profit_margin_pct",
        "pe_ratio": "pe_ratio",
        "pb_ratio": "pb_ratio",
        "dividend_yield": "dividend_yield_pct",
        "dividend_yield_pct": "dividend_yield_pct",
        "dividend_payout": "dividend_payout_ratio_pct",
        "dividend_payout_ratio_pct": "dividend_payout_ratio_pct",
        "interest_coverage": "interest_coverage",
        "market_cap": "market_cap_crore",
        "net_profit": "net_profit",
        "asset_turnover": "asset_turnover",
        "sales": "sales",
        "compounded_sales_growth": "revenue_cagr_5yr",
        "compounded_profit_growth": "pat_cagr_5yr",
        "return_on_equity_pct": "return_on_equity_pct",
        "free_cash_flow_cr": "free_cash_flow_cr",
    }